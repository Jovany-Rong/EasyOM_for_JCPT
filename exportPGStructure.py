#!/usr/local/bin python
#coding: utf-8

import pgx
from openpyxl import Workbook

pg = pgx.PgConnect('njcgknocoregis', 'njcgknocoregis', 'njcgknocoregis', '192.168.192.66', 30533)
tables = pg.tables('njcgknocoregis')

for table in tables:
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'name'
    ws['B1'] = 'type'
    ws['C1'] = 'is_not_null'
    ws['D1'] = 'comment'

    strucs = pg.table_structure(table)
    #print(struc)
    row = 1
    for struc in strucs:
        row += 1
        ws['A%s' % row] = struc[0]
        ws['B%s' % row] = struc[1]
        ws['C%s' % row] = struc[2]
        ws['D%s' % row] = struc[3]
    
    wb.save('table_structures/%s.xlsx' % table)