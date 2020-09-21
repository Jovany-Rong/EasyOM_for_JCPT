#!/usr/local/bin python
#coding: utf-8

import orax
from openpyxl import Workbook
import configparser as cp
import os

config = cp.ConfigParser()
config.read('export.conf', encoding='utf-8')

username = config.get('DEFAULT', 'username')
password = config.get('DEFAULT', 'password')
host = config.get('DEFAULT', 'host')
port = config.get('DEFAULT', 'port')
database = config.get('DEFAULT', 'database')
owners = config.get('ORACLE', 'owner')
ownerList = owners.split(',')

ora = orax.OraConnect(database, username, password, host, port)

for owner in ownerList:
    #owner = owner.strip.upper()
    print("dealing %s ..." % owner)
    tables = ora.tables(owner)
    if os.path.exists(owner) == False:
        os.makedirs(owner)

    for table in tables:
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'name'
        ws['B1'] = 'type'
        ws['C1'] = 'is_not_null'
        ws['D1'] = 'comment'

        strucs = ora.table_structure(table, owner)
        row = 1
        for struc in strucs:
            row += 1
            ws['A%s' % row] = struc[0]
            ws['B%s' % row] = struc[1]
            ws['C%s' % row] = struc[2]
            ws['D%s' % row] = struc[3]
    
        wb.save('%s/%s.xlsx' % (owner, table))
        print('\t%s/%s.xlsx done.' % (owner, table))

print("success.")
print("Powered by Chenfei Jovany Rong, GTMAP")