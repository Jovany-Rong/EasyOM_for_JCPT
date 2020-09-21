#!/usr/local/bin python
#-*-coding: utf-8-*-

import os
from openpyxl import load_workbook

def showXlsxFiles(filePath):
    ct = 0

    rList = []

    for root, dirs, files in os.walk(filePath):
        #global d

        if root == filePath:
            del root
            del dirs

            for file in files:
                if file.endswith(".xlsx"):
                    ct += 1

                    rList.append(file)
            
            break

    if ct == 0:
        print("所选路径中未找到xlsx文件！\n")

    else:
        print("共找到%s个xlsx文件。\n" % str(ct))

    return rList

def getSubDirs(filePath):
    for root, dirs, files in os.walk(filePath):
        if root == filePath:
            del files
            return dirs

def getSubFiles(filePath):
    res = []
    for root, dirs, files in os.walk(filePath):
        del root
        del dirs

        for file in files:
            res.append(file)
        
    return res

def isInCellRange(cellToCheck, cellRange):
    """
    to check a cell whether in a cell range
    :param cellToCheck:
    :param cellRange:
    :return:
        True : if cell in range
        False: if cell not in range
    """
    # logging.debug("cellToCheck=[%d:%d]", cellToCheck.row, cellToCheck.col_idx)
    # logging.debug("cellRange: row=[%d:%d] col=[%d:%d]",
    #              cellRange.min_row, cellRange.max_row, cellRange.min_col, cellRange.max_col)
    if (cellToCheck.row >= cellRange.min_row) and \
        (cellToCheck.row <= cellRange.max_row) and \
        (cellToCheck.col_idx >= cellRange.min_col) and \
        (cellToCheck.col_idx <= cellRange.max_col):

        return True
    else:
        return False


def getCellRangeValue(ws, cellRange):
    """
    get cell range value -&gt; the top left cell value
    :param cellRange:
    :return:
    """
    topLeftCell = ws.cell(row=cellRange.min_row, column=cellRange.min_col)
    topLeftCellValue = topLeftCell.value
    return topLeftCellValue

def getRealCellValue(ws, curCell):
    """
    for openpyxl, to get real value from row and column
    expecially for merged cell, will get its (same) value from top-left cell value

    :param row:
    :param column:
    :return:
    """

    realCellValue = curCell.value

    mergedCellsRangesList = ws.merged_cells.ranges
    # logging.info("mergedCellsRangesList=%s", mergedCellsRangesList)

    # Note:
    # to efficiency , we only check cell in range or not when its value is None
    # for all merged cell value is None
    if not realCellValue:
        for eachCellRange in mergedCellsRangesList:
            if isInCellRange(curCell, eachCellRange):
                cellRangeValue = getCellRangeValue(ws, eachCellRange)
                realCellValue =  cellRangeValue
                break

    return realCellValue

def rreplace(self, old, new, *max):
    count = len(self)
    if max and str(max[0]).isdigit():
        count = max[0]
    return new.join(self.rsplit(old, count))